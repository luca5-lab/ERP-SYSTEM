from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Cliente
from .forms import ClienteForm, ProveedorForm
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Proveedor
import openpyxl
from openpyxl.utils import get_column_letter
from usuarios.decorators import usuario_tipo_requerido
from usuarios.models import Proceso, Perfil, Venta, Compra, AccionHistorial,Factura, CotizacionServicio, CotizacionArticulo
from django.db.models import ProtectedError, Sum, Q
from django.contrib import messages
from django.core.paginator import Paginator


@login_required
@usuario_tipo_requerido('admin', 'administracion')
def informes(request):
    clientes_count = Cliente.objects.count()
    proveedores_count = Proveedor.objects.count()
    
    # --- VENTAS (Ingresos) ---
    resultado_ventas = Venta.objects.aggregate(total=Sum('monto_total'))
    total_ventas = resultado_ventas['total'] or 0 

    # --- COMPRAS (Egresos) ---
    # Sumamos el total_compra del modelo Compra
    resultado_compras = Compra.objects.aggregate(total=Sum('total_compra'))
    total_compras = resultado_compras['total'] or 0

    context = {
        'clientes_count': clientes_count,
        'proveedores_count': proveedores_count,
        'total_ventas': f"{total_ventas:,}".replace(",", "."),
        'total_compras': f"{total_compras:,}".replace(",", "."), # Formato CLP
    }
    return render(request, 'clientes/informes.html', context)

@login_required
@usuario_tipo_requerido('admin', 'taller', 'administracion')
def lista_proveedores(request):

    buscar = request.GET.get("buscar")

    proveedores = Proveedor.objects.all()

    if buscar:
        proveedores = proveedores.filter(rut__icontains=buscar)

    # PAGINADOR
    paginator = Paginator(proveedores, 5)  # 5 por página
    page = request.GET.get('page')
    proveedores = paginator.get_page(page)

    # Obtener el tipo de usuario del perfil
    tipo_usuario = getattr(request.user, 'perfil', None)
    tipo_usuario = tipo_usuario.tipo_usuario if tipo_usuario else None

    return render(
        request,
        'clientes/lista_proveedores.html',
        {
            'proveedores': proveedores,
            'buscar': buscar,
            'tipo_usuario': tipo_usuario
        }
    )

@login_required
@usuario_tipo_requerido('admin', 'administracion')
def registrar_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            prov = form.save()
            
            # --- REGISTRO EN HISTORIAL ---
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"PROVEEDORES: Registró nuevo proveedor - {prov.razon_social}",
                detalles=(
                    f"RUT: {prov.rut}\n"
                    f"Giro: {prov.giro}\n"
                    f"Contacto: {prov.telefono}\n"
                    f"Correo: {prov.correo}\n"
                    f"Dirección: {prov.direccion}"
                )
            )
            
            messages.success(request, f"Proveedor '{prov.razon_social}' registrado con éxito.")
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'clientes/registrar_proveedor.html', {'form': form})

@login_required
@usuario_tipo_requerido('admin', 'administracion')
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        # Capturamos los datos antiguos para el historial antes de guardar
        datos_anteriores = (
            f"Razón Social: {proveedor.razon_social}\n"
            f"RUT: {proveedor.rut}\n"
            f"Giro: {proveedor.giro}\n"
            f"Contacto: {proveedor.telefono}"
        )
        
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            # Guardamos los cambios
            proveedor_editado = form.save()
            
            # REGISTRO EN HISTORIAL
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"PROVEEDORES: Modificó datos de {proveedor_editado.razon_social}",
                detalles=(
                    f"DATOS ANTERIORES:\n{datos_anteriores}\n\n"
                    f"DATOS NUEVOS:\n"
                    f"Razón Social: {proveedor_editado.razon_social}\n"
                    f"RUT: {proveedor_editado.rut}\n"
                    f"Giro: {proveedor_editado.giro}\n"
                    f"Contacto: {proveedor_editado.telefono}"
                )
            )
            
            messages.success(request, f"Datos de '{proveedor.razon_social}' actualizados.")
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    
    return render(request, 'clientes/registrar_proveedor.html', {
        'form': form, 
        'editando': True,
        'proveedor': proveedor
    })

@login_required
@usuario_tipo_requerido('admin', 'administracion')
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        # Capturamos la info técnica antes de cualquier acción
        info_respaldo = (
            f"Razón Social: {proveedor.razon_social}\n"
            f"RUT: {proveedor.rut}\n"
            f"Giro: {proveedor.giro}"
        )
        
        try:
            nombre_prov = proveedor.razon_social
            proveedor.delete()
            
            # 1. REGISTRO DE ELIMINACIÓN EXITOSA
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"PROVEEDORES: Eliminó proveedor - {nombre_prov}",
                detalles=f"Registro borrado permanentemente.\nDatos de respaldo:\n{info_respaldo}"
            )
            
            messages.success(request, f"Proveedor '{nombre_prov}' eliminado con éxito.")
            return redirect('lista_proveedores')
            
        except ProtectedError:
            # 2. REGISTRO DE INTENTO FALLIDO (Auditoría de integridad)
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"SEGURIDAD: Intento de eliminación fallido - {proveedor.razon_social}",
                detalles=(
                    f"El usuario intentó eliminar un proveedor con facturas vinculadas.\n"
                    f"Sistema bloqueó la acción para preservar la integridad contable.\n"
                    f"{info_respaldo}"
                )
            )
            
            messages.error(request, 
                f"No se puede eliminar a '{proveedor.razon_social}' porque existen facturas de compras "
                "registradas a su nombre. El historial contable debe ser preservado."
            )
            return redirect('lista_proveedores')

    return render(request, 'clientes/confirmar_eliminar_proveedor.html', {'proveedor': proveedor})

@login_required
@usuario_tipo_requerido('admin', 'administracion') # Aseguramos que solo autorizados exporten
def exportar_clientes_excel(request):
    # 1. Obtener los objetos
    clientes_qs = Cliente.objects.all()
    cantidad_clientes = clientes_qs.count()

    # 2. Creamos la data manualmente
    data = []
    for c in clientes_qs:
        data.append({
            'RAZÓN SOCIAL': c.razon_social,
            'GIRO': c.giro,
            'RUT': c.rut,
            'DIRECCIÓN': c.direccion,
            'EMAIL': c.correo,
            'TELÉFONO': c.telefono,
            'FECHA REGISTRO': c.fecha_registro,
            'COND. PAGO': c.get_condiciones_pago_display() 
        })

    df = pd.DataFrame(data)

    # 3. Limpiar zona horaria
    if not df.empty:
        df['FECHA REGISTRO'] = df['FECHA REGISTRO'].dt.tz_localize(None)

    # --- REGISTRO EN HISTORIAL (Antes de enviar la respuesta) ---
    AccionHistorial.objects.create(
        usuario=request.user,
        accion="SEGURIDAD: Exportó Base de Clientes a Excel",
        detalles=(
            f"El usuario generó un reporte completo de clientes.\n"
            f"Cantidad de registros exportados: {cantidad_clientes}\n"
            f"Formato: XLSX (Excel)\n"
            f"Nota: Esta acción implica la descarga de datos sensibles de la empresa."
        )
    )

    # 4. Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Clientes_USBTech.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Base de Clientes')
        worksheet = writer.sheets['Base de Clientes']

        # --- ESTILOS ---
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border
            # Ajuste de ancho
            max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 5
            # Convertir numero de columna a letra (A, B, C...) de forma segura
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max_length

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")

    return response


@login_required
@usuario_tipo_requerido('admin', 'administracion', 'vendedor')
def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == 'POST':
        # 1. Capturamos los datos actuales antes de que el form los sobrescriba
        datos_previos = (
            f"Razón Social: {cliente.razon_social}\n"
            f"RUT: {cliente.rut}\n"
            f"Giro: {cliente.giro}\n"
            f"Contacto: {cliente.telefono}\n"
            f"Correo: {cliente.correo}"
        )
        
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            # 2. Guardamos los nuevos datos
            cliente_editado = form.save()
            
            # 3. REGISTRO EN HISTORIAL
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"CLIENTES: Editó ficha de {cliente_editado.razon_social}",
                detalles=(
                    f" DATOS ANTERIORES\n{datos_previos}\n\n"
                    f" DATOS NUEVOS\n"
                    f"Razón Social: {cliente_editado.razon_social}\n"
                    f"RUT: {cliente_editado.rut}\n"
                    f"Giro: {cliente_editado.giro}\n"
                    f"Contacto: {cliente_editado.telefono}\n"
                    f"Correo: {cliente_editado.correo}"
                )
            )
            
            messages.success(request, f"Los datos de '{cliente.razon_social}' han sido actualizados.")
            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=cliente)
    
    return render(request, 'clientes/registrar_cliente.html', {
        'form': form, 
        'editando': True,
        'cliente': cliente # Agregamos el objeto por si lo necesitas en el template
    })

@login_required
@usuario_tipo_requerido('admin', 'administracion', 'vendedor')
def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == 'POST':
        # Capturamos la información antes de cualquier acción para el historial
        info_respaldo = (
            f"Razón Social: {cliente.razon_social}\n"
            f"RUT: {cliente.rut}\n"
            f"Giro: {cliente.giro}"
        )
        
        try:
            nombre_cliente = cliente.razon_social
            cliente.delete()
            
            # 1. REGISTRO DE ELIMINACIÓN EXITOSA
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"CLIENTES: Eliminó cliente - {nombre_cliente}",
                detalles=f"Registro borrado permanentemente.\nDatos de respaldo:\n{info_respaldo}"
            )
            
            messages.success(request, f"El cliente '{nombre_cliente}' ha sido eliminado correctamente.")
            return redirect('lista_clientes')
            
        except ProtectedError:
            # 2. REGISTRO DE INTENTO FALLIDO (Auditoría de integridad)
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"SEGURIDAD: Intento de eliminación bloqueado - {cliente.razon_social}",
                detalles=(
                    f"El usuario intentó eliminar un cliente con historial transaccional (Ventas/Facturas).\n"
                    f"El sistema bloqueó la acción para no romper la integridad de los reportes.\n"
                    f"{info_respaldo}"
                )
            )
            
            messages.error(request, 
                f"ERROR DE INTEGRIDAD: No se puede eliminar a '{cliente.razon_social}' "
                "porque tiene historial de Ventas, Facturas o Cotizaciones. "
                "Para no romper la contabilidad, el sistema ha bloqueado el borrado."
            )
            return redirect('lista_clientes')

    return render(request, 'clientes/confirmar_eliminar.html', {'cliente': cliente})


@login_required
@usuario_tipo_requerido('admin', 'administracion', 'vendedor')
def registrar_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            
            # --- REGISTRO EN HISTORIAL ---
            AccionHistorial.objects.create(
                usuario=request.user,
                accion=f"CLIENTES: Registró nuevo cliente - {cliente.razon_social}",
                detalles=(
                    f"RUT: {cliente.rut}\n"
                    f"Giro: {cliente.giro}\n"
                    f"Correo: {cliente.correo}\n"
                    f"Teléfono: {cliente.telefono}\n"
                    f"Dirección: {cliente.direccion}"
                )
            )
            
            messages.success(request, f"Cliente '{cliente.razon_social}' registrado con éxito.")
            return redirect('lista_clientes')
    else:
        form = ClienteForm()
    return render(request, 'clientes/registrar_cliente.html', {'form': form})


@login_required
def lista_clientes(request):

    buscar = request.GET.get("buscar")

    clientes_list = Cliente.objects.all()

    if buscar:
        clientes_list = clientes_list.filter(rut__icontains=buscar)

    # PAGINADOR (5 registros)
    paginator = Paginator(clientes_list, 5)
    page_number = request.GET.get("page")
    clientes = paginator.get_page(page_number)

    tipo_usuario = getattr(request.user, 'perfil', None)
    tipo_usuario = tipo_usuario.tipo_usuario if tipo_usuario else None

    return render(
        request,
        'clientes/lista_clientes.html',
        {
            'clientes': clientes,
            'buscar': buscar,
            'tipo_usuario': tipo_usuario
        }
    )

@login_required
@usuario_tipo_requerido('admin', 'administracion')
def exportar_proveedores_excel(request):
    # 1. Obtener datos de proveedores
    proveedores = Proveedor.objects.all()
    cantidad_proveedores = proveedores.count()
    
    # 2. Crear la lista de diccionarios
    data = []
    for p in proveedores:
        data.append({
            'RAZÓN SOCIAL': p.razon_social,
            'GIRO': p.giro,
            'RUT': p.rut,
            'EMAIL': p.correo,
            'TELÉFONO': p.telefono or "",
            'COND. PAGO': p.get_condiciones_pago_display()
        })

    df = pd.DataFrame(data)

    # --- REGISTRO EN HISTORIAL DE SEGURIDAD ---
    AccionHistorial.objects.create(
        usuario=request.user,
        accion="SEGURIDAD: Exportó Base de Proveedores a Excel",
        detalles=(
            f"El usuario generó un reporte completo de proveedores.\n"
            f"Cantidad de proveedores exportados: {cantidad_proveedores}\n"
            f"Formato: XLSX"
        )
    )

    # 3. Configurar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Proveedores_USBTech.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Base de Proveedores')
        worksheet = writer.sheets['Base de Proveedores']

        # --- DEFINICIÓN DE ESTILOS ---
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # --- APLICAR ESTILOS Y AJUSTAR COLUMNAS ---
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border

            # Ancho automático robusto
            max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 5
            
            # Uso de la utilidad de openpyxl para evitar errores de letra
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = max_length

        # --- APLICAR ESTILOS A LOS DATOS ---
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=len(df) + 1,
            min_col=1,
            max_col=len(df.columns)
        ):
            for cell in row:
                cell.border = thin_border
                cell.alignment = alignment_left

    return response

@login_required
def lista_procesos_clientes(request):
    buscar = request.GET.get("buscar")
    
    # Obtenemos los procesos ordenados por los más recientes
    procesos = Proceso.objects.select_related('cliente', 'usuario').all().order_by('-creado')

    if buscar:
        # Filtramos por nombre de cliente o RUT
        procesos = procesos.filter(
            Q(cliente__razon_social__icontains=buscar) | 
            Q(cliente__rut__icontains=buscar)
        )

    return render(request, 'clientes/lista_procesos.html', {
        'procesos': procesos,
        'buscar': buscar
    })

@login_required
@usuario_tipo_requerido('admin', 'administracion')
def exportar_ventas_excel(request):
    # Traemos los datos base
    ventas_qs = Venta.objects.select_related(
        'cliente', 
        'factura'
    ).all().order_by('-fecha_registro')

    data = []
    for v in ventas_qs:
        n_cotizacion = "N/A"
        n_ot = "Sin OT"
        vendedor_cod = "N/A"

        if v.factura:
            f = v.factura
            
            # --- LIMPIEZA DE ID (Solución al error 7,7) ---
            id_numerico = None
            if f.origen_id:
                try:
                    # Convertimos a string, quitamos comas/puntos y pasamos a int
                    # Esto transforma "7,7" en 7
                    id_limpio = str(f.origen_id).split(',')[0].split('.')[0]
                    id_numerico = int(id_limpio)
                except (ValueError, TypeError):
                    id_numerico = None

            # 1. Buscar Cotización usando el ID limpio
            if id_numerico:
                if f.tipo == 'servicio':
                    cotiz = CotizacionServicio.objects.filter(pk=id_numerico).first()
                    if cotiz:
                        n_cotizacion = cotiz.get_numero_formateado()
                elif f.tipo == 'articulo':
                    # Aquí es donde fallaba porque pk apunta a n_presupuesto
                    cotiz = CotizacionArticulo.objects.filter(pk=id_numerico).first()
                    if cotiz:
                        n_cotizacion = cotiz.get_numero_formateado()
                        if cotiz.usuario and hasattr(cotiz.usuario, 'perfil'):
                            vendedor_cod = cotiz.usuario.perfil.codigo_vendedor or "Sin Código"

            # 2. Lógica de la OT
            if f.proceso_asociado:
                n_ot = f"OT-{str(f.proceso_asociado.id).zfill(5)}"
            elif f.tipo == 'servicio' and id_numerico:
                from .models import Proceso 
                proceso_oculto = Proceso.objects.filter(cotizacion_id=id_numerico).first()
                if proceso_oculto:
                    n_ot = f"OT-{str(proceso_oculto.id).zfill(5)}"
        else:
            n_ot = "Venta sin factura"

        data.append({
            'N° FACTURA': v.n_factura,
            'N° COTIZACIÓN': n_cotizacion,
            'VENDEDOR': vendedor_cod,
            'ORDEN TRABAJO': n_ot,         
            'CLIENTE': v.cliente.razon_social if v.cliente else "Sin Cliente",
            'RUT CLIENTE': v.cliente.rut if v.cliente else "",
            'FECHA REGISTRO': v.fecha_registro.date() if v.fecha_registro else "",
            'MONTO NETO ($)': v.monto_neto,
            'MONTO TOTAL ($)': v.monto_total
        })

    # ... (El resto del código de Pandas y ExcelWriter que ya tienes)
    # RECUERDA: En el ExcelWriter, usa las columnas ['H', 'I'] para el formato moneda.

    df = pd.DataFrame(data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Ventas_USBTech.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte de Ventas')
        worksheet = writer.sheets['Reporte de Ventas']

        # --- ESTILOS ---
        header_fill = PatternFill(start_color="0DCAF0", end_color="0DCAF0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border
            
            column_letter = get_column_letter(col_num)
            max_val = df[column_title].astype(str).map(len).max()
            adjusted_width = max(max_val, len(column_title)) + 4
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Estilo celdas de datos y formato moneda
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.border = thin_border
                # Ajustado a H e I que son los montos netos y totales
                if cell.column_letter in ['H', 'I']:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

    return response

@login_required
@usuario_tipo_requerido('admin', 'administracion')
def exportar_compras_excel(request):
    # 1. Obtener los datos de compras con el proveedor relacionado
    compras_qs = Compra.objects.select_related('proveedor').all().order_by('-fecha_registro')

    data = []
    for c in compras_qs:
        data.append({
            'N° FACTURA/BOLETA': c.n_factura if c.n_factura else "S/N",
            'PROVEEDOR': c.proveedor.razon_social if c.proveedor else "Sin Proveedor",
            'RUT PROVEEDOR': c.proveedor.rut if c.proveedor else "",
            'FECHA REGISTRO': c.fecha_registro, # Es un DateField, no necesita tzinfo=None
            'ESTADO PAGO': c.get_estado_pago_display(),
            'VALOR ABONADO ($)': c.valor_abonado,
            'SALDO PENDIENTE ($)': c.saldo_pendiente,
            'TOTAL COMPRA ($)': c.total_compra
        })

    # 2. Crear DataFrame
    df = pd.DataFrame(data)

    # 3. Configurar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Compras_USBTech.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte de Compras')
        worksheet = writer.sheets['Reporte de Compras']

        # --- ESTILOS ---
        # Usamos el color Danger (Rojo) para diferenciarlo de Ventas (Cyan)
        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid") 
        header_font = Font(color="FFFFFF", bold=True, size=12)
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Aplicar estilos a cabeceras y configurar ancho
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border
            
            # Cálculo de ancho automático
            col_letter = worksheet.cell(row=1, column=col_num).column_letter
            max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 5
            worksheet.column_dimensions[col_letter].width = max_length

        # Estilo celdas de datos
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")
                # Formato numérico para las columnas de dinero (F, G, H)
                if cell.column_letter in ['F', 'G', 'H']:
                    cell.number_format = '#,##0'

    return response